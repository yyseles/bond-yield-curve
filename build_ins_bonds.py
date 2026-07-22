#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_ins_bonds.py
以用户维护的 Excel《保险公司历年资本补充债发行情况.xlsx》为权威来源，
生成网页所需的 ins_bonds.json（同时含 资本补充债 与 永续债 两类）。

映射 (Excel data 表 -> ins_bonds schema)：
  发行人中文名称      -> issuer
  发行时债项评级      -> debtRating
  发行时主体评级      -> (用于 ratingStr 主体/债项)
  债券简称            -> bondShort
  证券全称            -> bondFull
  发行总额[亿元]      -> issueAmnt
  起息日期            -> valueDate / issueDate (按年分组)
  到期日期            -> mrtyDate (永续债即首个赎回日)
  债券期限(年)        -> bondPeriod
  票面利率(%)         -> couponRate
  note                -> status (存续/已赎回/已到期)
  bondType            -> 含"永续/无固定期限"为 永续债，否则 资本补充债
  industry            -> 由发行人名推断(产/寿/再保/集团)

生成的 bondDefinedCode 使用稳定合成键 XLSX-<行号>，避免与 chinamoney 真实
查询代码冲突；后续若用 chinamoney 抓取，可按 (issuer, bondShort) 去重合并。
"""
import json
import os
import re
from datetime import date

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(os.path.dirname(HERE), "保险公司历年资本补充债发行情况.xlsx")
OUT = os.path.join(HERE, "ins_bonds.json")

# 数据起始年份：2015/2016 年样本极少且口径不一，按需求自 2018 年起
START_YEAR = 2018


def infer_industry(issuer):
    if not issuer:
        return "其他"
    if "再保险" in issuer:
        return "再保"
    if re.search(r"财险|财产|产险", issuer):
        return "产险"
    if re.search(r"集团|控股", issuer):
        return "集团"
    return "寿险"


def to_float(v):
    if v in (None, "", "None"):
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def norm_date(v):
    if v is None:
        return ""
    s = str(v)
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            return ""
    return ""


def is_perp(bond_short, bond_full):
    s = f"{bond_short or ''} {bond_full or ''}"
    return ("永续" in s) or ("无固定期限" in s)


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["data"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr) if h}

    bonds = []
    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        issuer = row[idx.get("发行人中文名称")] if "发行人中文名称" in idx else None
        if issuer in (None, ""):
            continue
        issuer = str(issuer).strip()
        bond_short = str(row[idx.get("债券简称")] or "").strip()
        bond_full = str(row[idx.get("证券全称")] or "").strip()
        perp = is_perp(bond_short, bond_full)
        debt_rating = str(row[idx.get("发行时债项评级")] or "").strip()
        subj_rating = str(row[idx.get("发行时主体评级")] or "").strip()
        term = to_float(row[idx.get("债券期限(年)\n[单位] 年")])
        value_date = norm_date(row[idx.get("起息日期")])
        mrty_date = norm_date(row[idx.get("到期日期")])
        issue_amnt = to_float(row[idx.get("发行总额\n[单位] 亿元")])
        coupon = to_float(row[idx.get("票面利率(发行时)\n[单位] %")])
        status = str(row[idx.get("note")] or "存续").strip()
        if status not in ("存续", "已赎回", "已到期"):
            status = "存续"

        # 数据口径：仅保留 2018 年及以后发行的债
        issue_year = (value_date or "")[:4]
        if issue_year and issue_year.isdigit() and int(issue_year) < START_YEAR:
            continue

        if perp:
            bond_type = "永续债"
            bond_period = "5+N年"
            call_date = mrty_date
            exercise_flag = "是"
        else:
            bond_type = "资本补充债"
            bond_period = (f"{int(term)}年" if term else "")
            call_date = ""
            exercise_flag = "否"

        rec = {
            "bondDefinedCode": f"XLSX-{ri}",
            "issuer": issuer,
            "bondShort": bond_short,
            "bondFull": bond_full,
            "bondCode": "",
            "bondType": bond_type,
            "industry": infer_industry(issuer),
            "issueDate": value_date,
            "valueDate": value_date,
            "mrtyDate": mrty_date,
            "bondPeriod": bond_period,
            "planAmnt": issue_amnt,
            "issueAmnt": issue_amnt,
            "couponRate": coupon,
            "couponType": "附息式固定利率" if coupon is not None else "",
            "couponFrqncy": "年",
            "debtRating": debt_rating,
            "ratingStr": (f"{subj_rating}/{debt_rating}" if subj_rating else debt_rating),
            "exerciseFlag": exercise_flag,
            "callDate": call_date,
            "status": status,
            "source": "Excel(用户维护)",
        }
        bonds.append(rec)

    bonds.sort(key=lambda r: r.get("issueDate") or "", reverse=True)
    out = {
        "generatedAt": date.today().isoformat(),
        "source": "用户维护Excel《保险公司历年资本补充债发行情况.xlsx》+ 中国货币网(chinamoney)",
        "note": f"数据自 {START_YEAR} 年起（2015–2016 年样本较少，已剔除）",
        "count": len(bonds),
        "bonds": bonds,
    }
    json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

    from collections import Counter
    print(f"[done] 写出 {len(bonds)} 只 -> {OUT}")
    print("  按类型:", dict(Counter(b["bondType"] for b in bonds)))
    print("  按状态:", dict(Counter(b["status"] for b in bonds)))
    print("  按行业:", dict(Counter(b["industry"] for b in bonds)))
    # 逐年合计（与 Excel 汇总表核对）
    yr = {}
    for b in bonds:
        y = (b["issueDate"] or "")[:4]
        if not y:
            continue
        yr.setdefault(y, {"total": 0.0, "cap": 0.0, "perp": 0.0})
        yr[y]["total"] += b["issueAmnt"] or 0
        if b["bondType"] == "永续债":
            yr[y]["perp"] += b["issueAmnt"] or 0
        else:
            yr[y]["cap"] += b["issueAmnt"] or 0
    print("  逐年(总额/资本/永续)亿:")
    for y in sorted(yr):
        d = yr[y]
        print(f"    {y}: {round(d['total'],1)} / 资本 {round(d['cap'],1)} / 永续 {round(d['perp'],1)}")


if __name__ == "__main__":
    main()
