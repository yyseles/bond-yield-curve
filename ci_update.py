#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GitHub Actions CI 数据更新脚本
每天自动从中债网抓取最新国债即期利率，更新 data.json
数据源: https://yield.chinabond.com.cn/cbweb-mn/yc/bxjInit (csz=1, 即期利率)
通过 bxjDownload 接口下载完整 XLSX, 包含 0~50Y 每5天一个数据点

改进点:
- 使用北京时间判断"今天"，匹配中债网数据发布节奏
- 增加重试机制，应对中债网偶发性超时
- 仅抓取缺失的新日期，避免重复抓取已有数据（除非在回填窗口内）
- 详细的日志输出，便于排查

sub-1Y 短端曲线:
- bxjDownload 接口的完整 XLSX 已含 0~1Y 每5天一个点的密集曲线(共73个 sub-1Y 点)
- 之前只在 fetch 阶段提取整数年并丢弃了短端；ALM(现行) 规则需要 sub-1Y 月度精确采样
- 现额外维护 data_gov_spot_short_recent.json: 仅存近期窗口(默认365日)的 sub-1Y 短端,
  数据增量极小(约73点/日 × ~250日 ≈ 1.8万点, <0.3MB), 不动 data.json 全量历史
"""
import json
import os
import sys
import tempfile
import time
import argparse
from datetime import datetime, date, timedelta, timezone

import requests
from openpyxl import load_workbook

DATA_FILE = "data.json"
# sub-1Y 短端近期切片（供 ALM 规则的月度精确采样）
SHORT_FILE = "data_gov_spot_short_recent.json"
SHORT_WINDOW_DAYS = 365  # 与 RECENT_WINDOW_DAYS 对齐
CHINABOND_DOWNLOAD_URL = "https://yield.chinabond.com.cn/cbweb-mn/yc/bxjDownload"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://yield.chinabond.com.cn/cbweb-mn/yc/bxjInit?locale=zh_CN",
}

# 关键期限 1Y ~ 50Y (整数年)
ALL_TERMS = [f"{i}Y" for i in range(1, 51)]

# 北京时区 UTC+8
BJ_TZ = timezone(timedelta(hours=8))

MAX_RETRIES = 3
RETRY_DELAY = 5  # 秒

# ---------- 近期 12 个月滚动切片（供「利率曲线分析」板块快速首屏） ----------
# 切片 = 最后 365 个日历日（约 252 个交易日），体积恒定（~100KB/个），不随时间增长。
# 配合前端「先加载近期、后台加载全量」的渐进式加载，解决分析板块加载慢/白屏问题。
RECENT_MAP = {
    'data.json':         'data_gov_spot_recent.json',   # 国债即期全量 -> 近期切片（analysis 板块秒出用）
    'data_gov_ytm.json': 'data_gov_ytm_recent.json',
    'data_cdb.json':     'data_cdb_recent.json',
    'data_cdb_ytm.json': 'data_cdb_ytm_recent.json',
}
RECENT_WINDOW_DAYS = 365  # 约 12 个月


# ---------- 国开债即期 / 国债到期 / 国开债到期（中债 searchYc 接口） ----------
# 历史遗留：这三个全量文件此前只在手动回填时更新，每日 CI 从不维护，
# 导致它们会停在最后一次手动回填的日期（2026-07-24）。现并入每日 CI。
SEARCH_YC_URL = "https://yield.chinabond.com.cn/cbweb-mn/yc/searchYc"
SEARCH_YC_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://yield.chinabond.com.cn/cbweb-mn/yield_main?locale=zh_CN",
    "Content-Type": "application/x-www-form-urlencoded",
}
CDB_ID = "8a8b2ca037a7ca910137bfaa94fa5057"        # 国开债曲线（即期 qxll=1 / 到期 qxll=0）
GOV_YTM_ID = "2c9081e50a2f9606010a3068cae70001"    # 国债到期曲线
# (全量文件, 曲线ID, qxll, 名称) —— 需每日与国债即期日历同步补齐
SEARCHYC_CURVES = [
    ("data_cdb.json",     CDB_ID,     "1", "国开债即期"),
    ("data_gov_ytm.json", GOV_YTM_ID, "0", "国债到期"),
    ("data_cdb_ytm.json", CDB_ID,     "0", "国开债到期"),
]


def fetch_searchyc(curve_id, query_date, qxll):
    """从中债 searchYc 接口抓取指定曲线在指定日期的整数年(1~50Y)利率(%)。
    返回 {1: val, 2: val, ...} 或 None（无数据/失败）。带重试。"""
    params = {
        "xyzSelect": "txy", "workTimes": query_date, "dxbj": "0",
        "qxll": qxll, "yqqxN": "N", "yqqxK": "K",
        "ycDefIds": curve_id, "wrjxCBFlag": "0", "locale": "zh_CN",
    }
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.post(SEARCH_YC_URL, data=params, headers=SEARCH_YC_HEADERS, timeout=30)
            r.raise_for_status()
            data = r.json()
            if not data:
                if attempt < MAX_RETRIES:
                    print(f"  {query_date}: searchYc 空响应，第{attempt}次重试...")
                    time.sleep(RETRY_DELAY)
                    continue
                return None
            sd = data[0].get("seriesData", [])
            result = {}
            for tenor, val in sd:
                if abs(tenor - round(tenor)) < 1e-6 and 1 <= tenor <= 50:
                    result[int(tenor)] = round(float(val), 8)
            if len(result) >= 50:
                return result
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                continue
            return result if result else None
        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
            else:
                print(f"  {query_date}: searchYc 失败 - {e}")
                return None
    return None


def load_full_curve(file):
    """加载全量曲线文件（data_cdb.json / data_gov_ytm.json / data_cdb_ytm.json）"""
    if not os.path.exists(file):
        return {"dates": [], "terms": ALL_TERMS, "rows": []}
    with open(file, "r", encoding="utf-8") as f:
        d = json.load(f)
    if len(d.get("terms", [])) < 50:
        d["terms"] = ALL_TERMS
    return d


def update_searchyc_curves(ref_dates):
    """按国债即期交易日历，补齐各 searchYc 曲线的缺失日期（仅补齐，不重抓已有日）。"""
    if not ref_dates:
        return
    print("\n同步国开债即期 / 国债到期 / 国开债到期曲线（searchYc）：")
    for file, curve_id, qxll, name in SEARCHYC_CURVES:
        existing = load_full_curve(file)
        # 只向前补齐：曲线自身最后一日之后的交易日（不碰历史已有缺口，避免每日重复请求历史缺口）
        last = existing["dates"][-1] if existing["dates"] else None
        todo = [d for d in ref_dates if (last is None or d > last)]
        if not todo:
            print(f"  {name}: 已最新 ({existing['dates'][-1] if existing['dates'] else '无'})")
            continue
        d2r = {d: r for d, r in zip(existing["dates"], existing["rows"])}
        got = 0
        for d in todo:
            rates = fetch_searchyc(curve_id, d, qxll)
            if rates:
                d2r[d] = [rates.get(y) for y in range(1, 51)]
                got += 1
                print(f"  ✓ {name} {d}: 补齐")
            else:
                print(f"  - {name} {d}: 无数据，跳过（下次重试）")
            time.sleep(0.2)
        sd = sorted(d2r.keys())
        out = {"dates": sd, "terms": ALL_TERMS, "rows": [d2r[d] for d in sd]}
        with open(file, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False)
        print(f"  {name}: 补齐 {got}/{len(todo)} 天 → {file} (最新 {sd[-1]})")


def make_recent_slice(src, dst, window=RECENT_WINDOW_DAYS):
    if not os.path.exists(src):
        print(f"  跳过 {src}（不存在，未生成切片）")
        return False
    with open(src, 'r', encoding='utf-8') as f:
        d = json.load(f)
    dates = d.get('dates', [])
    if not dates:
        print(f"  跳过 {src}（无 dates）")
        return False
    last = datetime.strptime(dates[-1], '%Y-%m-%d')
    cutoff = (last - timedelta(days=window)).strftime('%Y-%m-%d')
    idx = [i for i, dt in enumerate(dates) if dt >= cutoff]
    if not idx:
        idx = list(range(len(dates)))
    out = {
        'source': d.get('source', ''),
        'note': f'近期 {window} 日滚动切片（自动生成，请勿手改）；全量见 {src}',
        'terms': d.get('terms', []),
        'dates': [dates[i] for i in idx],
        'rows': [d['rows'][i] for i in idx],
    }
    with open(dst, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False)
    print(f"  {src} ({len(dates)} 条) -> {dst} ({len(idx)} 条, {out['dates'][0]}~{out['dates'][-1]})")
    return True


def generate_recent_slices(base_dir='.'):
    print("\n生成近期切片（滚动窗口，体积恒定）：")
    for src, dst in RECENT_MAP.items():
        make_recent_slice(os.path.join(base_dir, src), os.path.join(base_dir, dst))


# ---------- 首页 4 卡片用的 summary.json ----------
# 4 曲线 → 各自的全量文件 → 显示名（与首页 dash-card 顺序保持一致）
SUMMARY_CURVES = [
    ('gov_spot', 'data.json',         '国债即期'),
    ('gov_ytm',  'data_gov_ytm.json', '国债到期'),
    ('cdb_spot', 'data_cdb.json',     '国开债即期'),
    ('cdb_ytm',  'data_cdb_ytm.json', '国开债到期'),
]
# 卡片上展示的关键期限：1Y 5Y 20Y 30Y + 10Y 单独作 hero（与 index.html 的 subTerms 一致）
SUMMARY_KEY_TERMS = ['1Y', '5Y', '10Y', '20Y', '30Y']
SUMMARY_FILE = 'summary.json'


def _summary_common_latest(base_dir):
    """取 4 曲线共同最新日期（min），保证任一卡片都不会显示未来日"""
    latest = None
    for _, fn, _ in SUMMARY_CURVES:
        p = os.path.join(base_dir, fn)
        if not os.path.exists(p):
            return None
        with open(p, 'r', encoding='utf-8') as f:
            d = json.load(f)
        ld = d['dates'][-1] if d.get('dates') else None
        if ld is None:
            return None
        latest = ld if latest is None or ld < latest else latest
    return latest


def _summary_build_one(filename, display_name, ref_date):
    with open(filename, 'r', encoding='utf-8') as f:
        d = json.load(f)
    dates, terms, rows = d['dates'], d['terms'], d['rows']
    if ref_date not in dates:
        return None
    ci = dates.index(ref_date)
    pi = ci - 1 if ci > 0 else None
    out_terms = {}
    for t in SUMMARY_KEY_TERMS:
        if t not in terms:
            out_terms[t] = {'value': None, 'change': None}
            continue
        ti = terms.index(t)
        cv = rows[ci][ti]
        if cv is None:
            out_terms[t] = {'value': None, 'change': None}
            continue
        pv = rows[pi][ti] if pi is not None else None
        chg = (cv - pv) if pv is not None else None
        out_terms[t] = {'value': cv, 'change': chg}
    return {'name': display_name, 'date': ref_date, 'terms': out_terms}


def generate_summary(base_dir='.'):
    """基于 4 曲线 data 文件生成 summary.json（首页 4 卡片数据源）。
    若任一曲线文件缺失或无 dates，跳过（不破坏 CI）。"""
    print("\n生成首页 4 卡片用的 summary.json：")
    ref = _summary_common_latest(base_dir)
    if ref is None:
        print("  跳过（某曲线文件缺失或无数据）")
        return False
    curves_out = {}
    for key, fn, name in SUMMARY_CURVES:
        c = _summary_build_one(os.path.join(base_dir, fn), name, ref)
        if c is None:
            print(f"  跳过（{fn} 中找不到 {ref}）")
            return False
        curves_out[key] = c
        hero = c['terms'].get('10Y') or {}
        v = hero.get('value')
        chg = hero.get('change')
        bp = f'{chg*100:+.1f}bp' if chg is not None else '—'
        print(f'  {key:<8} date={c["date"]}  10Y={v}  较上日 {bp}')
    payload = {
        'date': ref,
        'generatedAt': datetime.now(BJ_TZ).strftime('%Y-%m-%dT%H:%M:%S+08:00'),
        'curves': curves_out,
    }
    with open(os.path.join(base_dir, SUMMARY_FILE), 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, separators=(',', ':'))
    print(f"  写出 {SUMMARY_FILE} ({ref})")
    return True


def now_beijing() -> date:
    """返回北京时间今天的日期"""
    return datetime.now(BJ_TZ).date()


def fetch_spot_rates_chinabond(query_date: str):
    """
    从中债网 bxjDownload 接口下载 XLSX，提取即期利率。
    返回 (int_result, short_result):
      - int_result: {"1Y": rate, ..., "50Y": rate} (整数年, 百分比)
      - short_result: {0.0: rate, 0.0137: rate, ...} (sub-1Y 密集曲线, 年小数键, 百分比)
    任一部分无数据则返回 ({}, {})
    带重试机制。
    """
    params = {
        "gzr": query_date,
        "csz": "1",
        "locale": "zh_CN",
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.post(
                CHINABOND_DOWNLOAD_URL, params=params, headers=HEADERS, timeout=30
            )
            r.raise_for_status()

            # 检查响应是否为有效 Excel（中债网无数据时可能返回小体积非Excel内容）
            if len(r.content) < 200:
                if attempt < MAX_RETRIES:
                    print(f"  {query_date}: 响应过短({len(r.content)}B)，第{attempt}次重试...")
                    time.sleep(RETRY_DELAY)
                    continue
                print(f"  {query_date}: 无数据 (非交易日或未发布)")
                return {}, {}

            # 写入临时文件
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(r.content)
                tmp_path = tmp.name

            try:
                wb = load_workbook(tmp_path)
                ws = wb.active

                data = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    term_val = row[1]  # 标准期限(年)
                    rate_val = row[2]  # 平均值(%)
                    if term_val is not None and rate_val is not None:
                        data[float(term_val)] = float(rate_val)

                wb.close()
            finally:
                os.unlink(tmp_path)

            # 提取整数年
            int_result = {}
            for y in range(1, 51):
                val = data.get(float(y))
                if val is not None:
                    int_result[f"{y}Y"] = round(val, 8)

            # 提取 sub-1Y 短端密集曲线（年小数键，含 0Y）
            short_result = {}
            for t, v in data.items():
                if t < 1.0:
                    short_result[round(t, 6)] = round(v, 8)

            if not int_result and not short_result:
                print(f"  {query_date}: 无数据 (非交易日或未发布)")
                return {}, {}

            return int_result, short_result

        except Exception as e:
            if attempt < MAX_RETRIES:
                print(f"  {query_date}: 请求失败({e})，第{attempt}次重试...")
                time.sleep(RETRY_DELAY)
            else:
                print(f"  {query_date}: 请求失败 - {e}")
                return {}, {}

    return {}, {}


def load_existing_data() -> dict:
    """加载现有 data.json"""
    if not os.path.exists(DATA_FILE):
        return {"dates": [], "terms": ALL_TERMS, "rows": []}

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    if len(data.get("terms", [])) < 50:
        data["terms"] = ALL_TERMS

    return data


def save_data(data: dict):
    """保存 data.json"""
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def load_existing_short() -> dict:
    """加载现有 data_gov_spot_short_recent.json（sub-1Y 短端近期切片）"""
    if not os.path.exists(SHORT_FILE):
        return {"terms": [], "dates": [], "rows": []}
    with open(SHORT_FILE, "r", encoding="utf-8") as f:
        d = json.load(f)
    if not isinstance(d, dict) or "dates" not in d:
        return {"terms": [], "dates": [], "rows": []}
    return d


def save_short(data: dict):
    with open(SHORT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def get_recent_dates(all_dates, window=SHORT_WINDOW_DAYS):
    """返回 all_dates 中落在最近 window 天内的交易日（升序）"""
    if not all_dates:
        return []
    last = datetime.strptime(all_dates[-1], '%Y-%m-%d')
    cutoff = (last - timedelta(days=window)).strftime('%Y-%m-%d')
    return [d for d in all_dates if d >= cutoff]


def maintain_short_file(existing, short_new=None, verbose=True):
    """
    维护 data_gov_spot_short_recent.json：
    - 取 existing(全量 data.json) 的近期窗口交易日
    - 合并本次新抓的 short_new（date->short_dict）
    - 对缺失日从 bxjDownload 回填短端
    - 写出近期窗口的 short 文件（terms=sub-1Y 年小数列表, dates, rows）
    不修改 data.json。
    """
    if short_new is None:
        short_new = {}
    recent = get_recent_dates(existing.get("dates", []))
    if not recent:
        if verbose:
            print("short: 无近期交易日，跳过")
        return

    short_store = load_existing_short()
    short_terms = short_store.get("terms")
    if not short_terms:
        # 从本次新抓样本或抓近期最后一日确定 tenor 列表
        sample = next(iter(short_new.values()), None)
        if sample is None:
            if verbose:
                print("short: 无现有 tenor，抓取近期最后一日确定 sub-1Y 锚点...")
            s_int, s_short = fetch_spot_rates_chinabond(recent[-1])
            sample = s_short
        if sample:
            short_terms = sorted(sample.keys())

    if not short_terms:
        if verbose:
            print("short: 仍无法确定 tenor，跳过")
        return

    # date -> row
    d2r = {}
    for i, d in enumerate(short_store.get("dates", [])):
        d2r[d] = short_store["rows"][i]

    # 合并本次新抓
    for d, sh in short_new.items():
        if d in recent:
            d2r[d] = [sh.get(t) for t in short_terms]

    # 回填缺失日
    missing = [d for d in recent if d not in d2r]
    backfilled = 0
    if missing:
        if verbose:
            print(f"short: 回填 {len(missing)} 个缺失日 (近期窗口 {recent[0]}~{recent[-1]})...")
        for i, d in enumerate(missing):
            s_int, s_short = fetch_spot_rates_chinabond(d)
            if s_short:
                d2r[d] = [s_short.get(t) for t in short_terms]
                backfilled += 1
            if verbose and (i + 1) % 25 == 0:
                print(f"  short 回填进度 {i+1}/{len(missing)} (已成功 {backfilled})")
            time.sleep(0.15)  # 轻量限速，避免触发中债网限流

    # 仅保留近期窗口
    dates = [d for d in recent if d in d2r]
    rows = [d2r[d] for d in dates]
    save_short({"terms": short_terms, "dates": dates, "rows": rows})
    if verbose:
        print(f"short: 写出 {SHORT_FILE} ({len(dates)} 日, {len(short_terms)} 个 sub-1Y 期限, 回填成功 {backfilled}/{len(missing)})")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--short-only', action='store_true',
                        help='仅维护 sub-1Y 短端文件，不抓取/更新 data.json 整数年')
    args = parser.parse_args()

    print("=" * 55)
    print("  国债即期利率 · CI 自动更新 (中债网)")
    print(f"  北京时间: {datetime.now(BJ_TZ).strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)

    existing = load_existing_data()
    print(f"现有数据: {len(existing['dates'])} 条")

    if args.short_only:
        print("模式: 仅维护 sub-1Y 短端文件")
        maintain_short_file(existing, verbose=True)
        generate_recent_slices()
        generate_summary()
        print("\n✅ short-only 完成")
        return

    today_bj = now_beijing()
    today_str = today_bj.strftime("%Y-%m-%d")

    # 确定起始日期
    if existing["dates"]:
        last_date = existing["dates"][-1]
        # 从最后日期的次日开始抓新数据
        fetch_start_new = (
            datetime.strptime(last_date, "%Y-%m-%d") + timedelta(days=1)
        ).strftime("%Y-%m-%d")
    else:
        fetch_start_new = "2020-01-01"

    # 只抓最后日期之后的新数据（回填已完成，不再重复抓取）
    fetch_start = fetch_start_new
    print(f"抓取范围: {fetch_start} → {today_str}")
    print(f"  - 已有最后日期: {existing['dates'][-1] if existing['dates'] else '无'}")

    # 逐日抓取
    all_new = {}
    short_new = {}
    current = datetime.strptime(fetch_start, "%Y-%m-%d")
    end = datetime.strptime(today_str, "%Y-%m-%d")

    skipped = 0
    fetched = 0
    while current <= end:
        ds = current.strftime("%Y-%m-%d")
        # 跳过周末
        if current.weekday() < 5:
            rates, short = fetch_spot_rates_chinabond(ds)
            if rates:
                all_new[ds] = rates
                short_new[ds] = short
                fetched += 1
                print(f"  ✓ {ds}: {len(rates)} 个整数年期限, sub-1Y {len(short)} 点")
            else:
                skipped += 1
        current += timedelta(days=1)

    print(f"\n获取: {fetched} 个交易日, 跳过/无数据: {skipped} 天")

    if not all_new:
        print("\n⚠ 没有获取到新数据（可能当日数据尚未发布或非交易日），跳过 data.json 更新")
    else:
        # 合并数据
        date_to_row = {}
        for i, d in enumerate(existing["dates"]):
            date_to_row[d] = existing["rows"][i]

        new_count = 0
        update_count = 0
        for d in sorted(all_new.keys()):
            rates = all_new[d]
            row = [rates.get(t) for t in ALL_TERMS]
            if d in date_to_row:
                date_to_row[d] = row
                update_count += 1
            else:
                date_to_row[d] = row
                new_count += 1

        sorted_dates = sorted(date_to_row.keys())
        sorted_rows = [date_to_row[d] for d in sorted_dates]

        output = {"dates": sorted_dates, "terms": ALL_TERMS, "rows": sorted_rows}
        save_data(output)

        print(f"\n✅ 更新完成: 新增 {new_count} 条, 修正 {update_count} 条")
        print(f"   总计: {len(sorted_dates)} 条, {sorted_dates[0]} ~ {sorted_dates[-1]}")

    # 维护 sub-1Y 短端文件（合并本次新抓 + 回填缺失）
    maintain_short_file(existing, short_new, verbose=True)

    # 同步国开债即期 / 国债到期 / 国开债到期（按国债即期日历补齐缺失日）
    update_searchyc_curves(existing.get("dates", []))

    # 重新生成近期切片（滚动窗口，体积恒定 ~100KB/个，供分析板块快速首屏）
    generate_recent_slices()

    # 生成首页 4 卡片用的 summary.json（gov_spot / gov_ytm / cdb_spot / cdb_ytm 各卡片）
    generate_summary()


if __name__ == "__main__":
    main()
