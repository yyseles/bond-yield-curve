#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GitHub Actions CI 数据更新脚本
每天自动从中债网抓取最新国债即期利率，更新 data.json
数据源: https://yield.chinabond.com.cn/cbweb-mn/yc/bxjInit (csz=1, 即期利率)
通过 bxjDownload 接口下载完整 XLSX, 包含 0~50Y 每5天一个数据点

改进点:
- 使用北京时间判断"今天"，匹配中债网数据发布节奏
- 增加重试机制，应对中债网偶发性超时
- 仅抓取缺失的新日期，避免重复抓取已有数据（除非在回填窗口内）
- 详细的日志输出，便于排查
"""
import json
import os
import sys
import tempfile
import time
from datetime import datetime, date, timedelta, timezone

import requests
from openpyxl import load_workbook

DATA_FILE = "data.json"
CHINABOND_DOWNLOAD_URL = "https://yield.chinabond.com.cn/cbweb-mn/yc/bxjDownload"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://yield.chinabond.com.cn/cbweb-mn/yc/bxjInit?locale=zh_CN",
}

# 关键期限 1Y ~ 50Y (整数年)
ALL_TERMS = [f"{i}Y" for i in range(1, 51)]

# 北京时区 UTC+8
BJ_TZ = timezone(timedelta(hours=8))

MAX_RETRIES = 3
RETRY_DELAY = 5  # 秒

# ---------- 近期 12 个月滚动切片（供「利率曲线分析」板块快速首屏） ----------
# 切片 = 最后 365 个日历日（约 252 个交易日），体积恒定（~100KB/个），不随时间增长。
# 配合前端「先加载近期、后台加载全量」的渐进式加载，解决分析板块加载慢/白屏问题。
RECENT_MAP = {
    'data.json':         'data_gov_spot_recent.json',
    'data_gov_ytm.json': 'data_gov_ytm_recent.json',
    'data_cdb.json':     'data_cdb_recent.json',
    'data_cdb_ytm.json': 'data_cdb_ytm_recent.json',
}
RECENT_WINDOW_DAYS = 365  # 约 12 个月


def make_recent_slice(src, dst, window=RECENT_WINDOW_DAYS):
    if not os.path.exists(src):
        print(f"  跳过 {src}（不存在，未生成切片）")
        return False
    with open(src, 'r', encoding='utf-8') as f:
        d = json.load(f)
    dates = d.get('dates', [])
    if not dates:
        print(f"  跳过 {src}（无 dates）")
        return False
    last = datetime.strptime(dates[-1], '%Y-%m-%d')
    cutoff = (last - timedelta(days=window)).strftime('%Y-%m-%d')
    idx = [i for i, dt in enumerate(dates) if dt >= cutoff]
    if not idx:
        idx = list(range(len(dates)))
    out = {
        'source': d.get('source', ''),
        'note': f'近期 {window} 日滚动切片（自动生成，请勿手改）；全量见 {src}',
        'terms': d.get('terms', []),
        'dates': [dates[i] for i in idx],
        'rows': [d['rows'][i] for i in idx],
    }
    with open(dst, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False)
    print(f"  {src} ({len(dates)} 条) -> {dst} ({len(idx)} 条, {out['dates'][0]}~{out['dates'][-1]})")
    return True


def generate_recent_slices(base_dir='.'):
    print("\n生成近期切片（滚动窗口，体积恒定）：")
    for src, dst in RECENT_MAP.items():
        make_recent_slice(os.path.join(base_dir, src), os.path.join(base_dir, dst))


def now_beijing() -> date:
    """返回北京时间今天的日期"""
    return datetime.now(BJ_TZ).date()


def fetch_spot_rates_chinabond(query_date: str) -> dict:
    """
    从中债网 bxjDownload 接口下载 XLSX，提取整数年即期利率。
    返回 {"1Y": rate, "2Y": rate, ... "50Y": rate} 或空字典（非交易日/数据未发布）
    带重试机制。
    """
    params = {
        "gzr": query_date,
        "csz": "1",
        "locale": "zh_CN",
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.post(
                CHINABOND_DOWNLOAD_URL, params=params, headers=HEADERS, timeout=30
            )
            r.raise_for_status()

            # 检查响应是否为有效 Excel（中债网无数据时可能返回小体积非Excel内容）
            if len(r.content) < 200:
                if attempt < MAX_RETRIES:
                    print(f"  {query_date}: 响应过短({len(r.content)}B)，第{attempt}次重试...")
                    time.sleep(RETRY_DELAY)
                    continue
                print(f"  {query_date}: 无数据 (非交易日或未发布)")
                return {}

            # 写入临时文件
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(r.content)
                tmp_path = tmp.name

            try:
                wb = load_workbook(tmp_path)
                ws = wb.active

                data = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    term_val = row[1]  # 标准期限(年)
                    rate_val = row[2]  # 平均值(%)
                    if term_val is not None and rate_val is not None:
                        data[float(term_val)] = float(rate_val)

                wb.close()
            finally:
                os.unlink(tmp_path)

            # 提取整数年
            result = {}
            for y in range(1, 51):
                val = data.get(float(y))
                if val is not None:
                    result[f"{y}Y"] = round(val, 8)

            if not result:
                print(f"  {query_date}: 无数据 (非交易日或未发布)")
                return {}

            return result

        except Exception as e:
            if attempt < MAX_RETRIES:
                print(f"  {query_date}: 请求失败({e})，第{attempt}次重试...")
                time.sleep(RETRY_DELAY)
            else:
                print(f"  {query_date}: 请求失败 - {e}")
                return {}

    return {}


def load_existing_data() -> dict:
    """加载现有 data.json"""
    if not os.path.exists(DATA_FILE):
        return {"dates": [], "terms": ALL_TERMS, "rows": []}

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    if len(data.get("terms", [])) < 50:
        data["terms"] = ALL_TERMS

    return data


def save_data(data: dict):
    """保存 data.json"""
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def main():
    print("=" * 55)
    print("  国债即期利率 · CI 自动更新 (中债网)")
    print(f"  北京时间: {datetime.now(BJ_TZ).strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)

    existing = load_existing_data()
    print(f"现有数据: {len(existing['dates'])} 条")

    today_bj = now_beijing()
    today_str = today_bj.strftime("%Y-%m-%d")

    # 确定起始日期
    if existing["dates"]:
        last_date = existing["dates"][-1]
        # 从最后日期的次日开始抓新数据
        fetch_start_new = (
            datetime.strptime(last_date, "%Y-%m-%d") + timedelta(days=1)
        ).strftime("%Y-%m-%d")
    else:
        fetch_start_new = "2020-01-01"

    # 只抓最后日期之后的新数据（回填已完成，不再重复抓取）
    fetch_start = fetch_start_new
    print(f"抓取范围: {fetch_start} → {today_str}")
    print(f"  - 已有最后日期: {existing['dates'][-1] if existing['dates'] else '无'}")

    # 逐日抓取
    all_new = {}
    current = datetime.strptime(fetch_start, "%Y-%m-%d")
    end = datetime.strptime(today_str, "%Y-%m-%d")

    skipped = 0
    fetched = 0
    while current <= end:
        ds = current.strftime("%Y-%m-%d")
        # 跳过周末
        if current.weekday() < 5:
            rates = fetch_spot_rates_chinabond(ds)
            if rates:
                all_new[ds] = rates
                fetched += 1
                print(f"  ✓ {ds}: {len(rates)} 个期限")
            else:
                skipped += 1
        current += timedelta(days=1)

    print(f"\n获取: {fetched} 个交易日, 跳过/无数据: {skipped} 天")

    if not all_new:
        print("\n⚠ 没有获取到新数据（可能当日数据尚未发布或非交易日），跳过 data.json 更新")
    else:
        # 合并数据
        date_to_row = {}
        for i, d in enumerate(existing["dates"]):
            date_to_row[d] = existing["rows"][i]

        new_count = 0
        update_count = 0
        for d in sorted(all_new.keys()):
            rates = all_new[d]
            row = [rates.get(t) for t in ALL_TERMS]
            if d in date_to_row:
                date_to_row[d] = row
                update_count += 1
            else:
                date_to_row[d] = row
                new_count += 1

        sorted_dates = sorted(date_to_row.keys())
        sorted_rows = [date_to_row[d] for d in sorted_dates]

        output = {"dates": sorted_dates, "terms": ALL_TERMS, "rows": sorted_rows}
        save_data(output)

        print(f"\n✅ 更新完成: 新增 {new_count} 条, 修正 {update_count} 条")
        print(f"   总计: {len(sorted_dates)} 条, {sorted_dates[0]} ~ {sorted_dates[-1]}")

    # 重新生成近期切片（滚动窗口，体积恒定 ~100KB/个，供分析板块快速首屏）
    generate_recent_slices()


if __name__ == "__main__":
    main()
